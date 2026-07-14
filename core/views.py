from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from customer.models import Customer
from django.contrib import messages
from accounts.models import User
from django.db.models import Q, Sum
from datetime import datetime, date, timedelta
from django.utils import timezone
import re
from .models import Billing, Payment, Notification
from decimal import Decimal
from openpyxl import load_workbook
from collections import defaultdict
from django.contrib.auth.hashers import make_password
from django.db import transaction
from django.db.models import Sum, Count
from django.db.models.functions import Coalesce
import json

@login_required
def admin_dashboard(request):

    # ==========================
    # Dashboard Cards
    # ==========================
    total_connections = Customer.objects.count()

    pending_applicants = Customer.objects.filter(
        status="new"
    ).count()

    current_month = date.today().strftime("%B %Y")

    total_bills = Billing.objects.count()

    paid_bills = Billing.objects.filter(
        status="paid"
    ).count()

    if total_bills > 0:
        collection_rate = round(
            (paid_bills / total_bills) * 100,
            1
        )
    else:
        collection_rate = 0

    # ==========================
    # Recent Billings
    # ==========================
    recent_billings = (
        Billing.objects
        .select_related("customer")
        .order_by("-created_at")[:10]
    )

    # ==========================
    # Consumption Per Barangay
    # ==========================
    barangay_totals = defaultdict(float)

    billings = Billing.objects.select_related("customer")

    for bill in billings:

        address = bill.customer.address or ""

        barangay = "Unknown"

        # Address Format:
        # Purok 1, Barangay Caratagan, Pio Duran, Albay
        # Purok 2, Barangay 1, Pio Duran, Albay

        parts = [part.strip() for part in address.split(",")]

        for part in parts:

            if part.lower().startswith("barangay"):
                barangay = part
                break

        barangay_totals[barangay] += float(bill.consumption)

    barangay_consumption = [
        {
            "barangay": barangay,
            "total_consumption": total
        }
        for barangay, total in sorted(barangay_totals.items())
    ]

    context = {
        "total_connections": total_connections,
        "pending_applicants": pending_applicants,
        "current_month": current_month,
        "collection_rate": collection_rate,
        "recent_billings": recent_billings,
        "barangay_consumption": barangay_consumption,
    }

    return render(
        request,
        "admin/admin_dashboard.html",
        context
    )

@login_required
def customer_list(request):
    customers = Customer.objects.all()

    return render(request, 'admin/customers.html', {
        'customers': customers
    })

@login_required
def add_customer(request):

    if request.method == "POST":

        user = User.objects.create_user(
            username=request.POST.get('username'),
            password=request.POST.get('password'),
            role='customer'  
        )

        Customer.objects.create(
            user=user,
            fullname=request.POST.get('fullname'),
            submitter_no=request.POST.get('submitter_no'),
            address=request.POST.get('address'),
        )

        return redirect('customers')

    return render(request, 'admin/add_customer.html')

@login_required
def import_customers(request):

    if request.method != "POST":
        return redirect("customers")

    excel_file = request.FILES.get("excel_file")

    if not excel_file:
        messages.error(request, "Please upload an Excel file.")
        return redirect("customers")

    try:
        workbook = load_workbook(
            excel_file,
            read_only=True,
            data_only=True
        )
    except Exception:
        messages.error(request, "Invalid Excel file.")
        return redirect("customers")

    sheet = workbook.active

    headers = [
        str(h).strip().lower() if h else ""
        for h in next(sheet.iter_rows(max_row=1, values_only=True))
    ]

    required_headers = [
        "fullname",
        "submitter no.",
        "address",
    ]

    for header in required_headers:
        if header not in headers:
            messages.error(request, f"Missing column: {header}")
            return redirect("customers")

    fullname_col = headers.index("fullname")
    submitter_col = headers.index("submitter no.")
    address_col = headers.index("address")

    existing_submitters = set(
        Customer.objects.values_list("submitter_no", flat=True)
    )

    existing_usernames = set(
        User.objects.values_list("username", flat=True)
    )

    users = []
    customer_data = []

    imported = 0
    skipped = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):

        fullname = str(row[fullname_col] or "").strip()
        submitter_no = str(row[submitter_col] or "").strip()
        address = str(row[address_col] or "").strip()

        if not fullname and not submitter_no and not address:
            continue

        if submitter_no in existing_submitters:
            skipped += 1
            continue

        username = submitter_no

        counter = 1
        while username in existing_usernames:
            username = f"{submitter_no}_{counter}"
            counter += 1

        existing_usernames.add(username)
        existing_submitters.add(submitter_no)

        users.append(
            User(
                username=username,
                password=make_password(submitter_no),
                role="customer",
            )
        )

        customer_data.append({
            "username": username,
            "fullname": fullname,
            "submitter_no": submitter_no,
            "address": address,
        })

    try:
        with transaction.atomic():

            # Insert all users at once
            User.objects.bulk_create(users, batch_size=500)

            # Fetch inserted users
            user_map = {
                user.username: user
                for user in User.objects.filter(
                    username__in=[u.username for u in users]
                )
            }

            customers = []

            for item in customer_data:

                customers.append(
                    Customer(
                        user=user_map[item["username"]],
                        fullname=item["fullname"],
                        submitter_no=item["submitter_no"],
                        address=item["address"],
                        status="old",
                    )
                )

            # Insert all customers at once
            Customer.objects.bulk_create(customers, batch_size=500)

            imported = len(customers)

    except Exception as e:
        messages.error(request, f"Import failed: {e}")
        return redirect("customers")

    workbook.close()

    messages.success(
        request,
        f"{imported} customer(s) imported successfully. "
        f"{skipped} duplicate(s) skipped."
    )

    return redirect("customers")

# @login_required
# def approve_applicant(request, id):
#     if request.method == "POST":
#         customer = get_object_or_404(Customer, id=id)

#         if not customer.account_number:
#             year = datetime.now().year

#             last_customer = Customer.objects.filter(
#                 account_number__startswith=f"PWSS-{year}"
#             ).order_by("-account_number").first()

#             if last_customer:
#                 last_number = int(last_customer.account_number.split("-")[-1])
#                 next_number = last_number + 1
#             else:
#                 next_number = 1

#             customer.account_number = f"PWSS-{year}-{next_number:04d}"

#         customer.status = "active"
#         customer.is_active = True

#         customer.user.is_active = True
#         customer.user.save()

#         customer.save()

#         messages.success(request, "Applicant activated successfully.")

#     return redirect("new_applicants")


# @login_required
# def decline_applicant(request, id):
#     if request.method == "POST":
#         customer = get_object_or_404(Customer, id=id)

#         customer.status = "decline"
#         customer.is_active = False
#         customer.save()

#         customer.user.is_active = False
#         customer.user.save()

#         messages.success(request, "Applicant declined.")

#     return redirect("new_applicants")

# @login_required
# def suspend_customer(request, id):
#     if request.method == "POST":
#         customer = get_object_or_404(Customer, id=id)

#         customer.status = "inactive"
#         customer.is_active = False
#         customer.save()

#         customer.user.is_active = False
#         customer.user.save()

#         messages.success(request, f"{customer.fullname} has been suspended.")

#     return redirect("customers")


@login_required
def delete_customer(request, id):
    if request.method == "POST":
        customer = get_object_or_404(Customer, id=id)

        # Delete the Django user.
        # Because Customer has OneToOneField(User, on_delete=models.CASCADE),
        # deleting the user automatically deletes the Customer profile.
        customer.user.delete()

        messages.success(request, "Customer account deleted successfully.")

    return redirect("customers")


@login_required
def customer_profile(request, customer_id):

    customer = get_object_or_404(Customer, pk=customer_id)

    billings = Billing.objects.filter(
        customer=customer
    ).order_by("-billing_month")

    payments = Payment.objects.select_related(
        "billing"
    ).filter(
        billing__customer=customer
    ).order_by("-payment_date")

    total_consumption = billings.aggregate(
        total=Sum("consumption")
    )["total"] or Decimal("0.00")

    total_billed = billings.aggregate(
        total=Sum("total_amount")
    )["total"] or Decimal("0.00")

    total_paid = payments.aggregate(
        total=Sum("amount_paid")
    )["total"] or Decimal("0.00")

    return render(request, "admin/customer_profile.html", {
        "customer": customer,
        "billings": billings,
        "payments": payments,
        "total_consumption": total_consumption,
        "total_billed": total_billed,
        "total_paid": total_paid,
    })

# @login_required
# def reactivate_customer(request, id):
#     if request.method == "POST":
#         customer = get_object_or_404(Customer, id=id)

#         customer.status = "active"
#         customer.is_active = True
#         customer.save()

#         customer.user.is_active = True
#         customer.user.save()

#         messages.success(request, f"{customer.fullname} has been reactivated.")

#     return redirect("customers")

@login_required
def billing(request):
    billings = Billing.objects.select_related("customer").order_by(
        "-billing_month",
        "-created_at"
    )

    return render(request, "admin/billing.html", {
        "billings": billings
    })

@login_required
def create_bill(request):

    if request.method == "POST":

        customer = get_object_or_404(
            Customer,
            id=request.POST.get("customer")
        )

        billing_month = request.POST.get("billing_month")
        due_date = request.POST.get("due_date")

        previous_reading = Decimal(request.POST.get("previous_reading") or "0.00" ) 
        current_reading = Decimal(request.POST.get("current_reading") or "0.00")

        connection_fee = Decimal(request.POST.get("connection_fee") or "0.00")
        reconnection_fee = Decimal(request.POST.get("reconnection_fee") or "0.00")
        violation_fee = Decimal(request.POST.get("violation_fee") or "0.00")

        # Prevent duplicate bill
        if Billing.objects.filter(
            customer=customer,
            billing_month=billing_month
        ).exists():
            messages.error(
                request,
                "This customer already has a bill for the selected billing month."
            )
            return redirect("create_bill")

        # Validate reading
        if current_reading < previous_reading:
            messages.error(
                request,
                "Current reading cannot be less than the previous reading."
            )
            return redirect("create_bill")

        # Create bill
        Billing.objects.create(
            customer=customer,
            billing_month=billing_month,
            previous_reading=previous_reading,
            current_reading=current_reading,
            due_date=due_date,
            rate_per_cubic=Decimal(request.POST.get("rate_per_cubic")),

            connection_fee=connection_fee,
            reconnection_fee=reconnection_fee,
            violation_fee=violation_fee,

            status="unpaid",
        )

        # Change status after first billing
        if customer.status == "new":
            customer.status = "old"
            customer.save(update_fields=["status"])

        messages.success(request, "Water bill created successfully.")
        return redirect("create_bill")

    customers = Customer.objects.all()

    customer_data = []

    for customer in customers:

        last_bill = Billing.objects.filter(
            customer=customer
        ).order_by("-billing_month").first()

        previous = last_bill.current_reading if last_bill else Decimal("0.00")

        customer_data.append({
            "customer": customer,
            "previous": previous,
            "is_new": customer.status == "new",
        })

    return render(request, "admin/create_bill.html", {
        "customer_data": customer_data
    })

@login_required
def payment(request):

    today = timezone.localdate()

    billings = Billing.objects.select_related(
        "customer"
    ).filter(
        status="unpaid"
    ).order_by(
        "customer__fullname",
        "-billing_month"
    )

    for bill in billings:

        # Check if overdue and penalty not yet applied
        if bill.due_date < today and (bill.penalty_fee or Decimal("0.00")) == Decimal("0.00"):

            # 10% penalty based on the rate_per_cubic
            bill.penalty_fee = bill.rate_per_cubic * Decimal("0.10")

            # penalty for total_amount
            # bill.penalty_fee = bill.total_amount * Decimal("0.10")

            # Save the bill (save() will automatically recalculate total_amount)
            bill.save()

    return render(request, "admin/payments.html", {
        "billings": billings
    })

@login_required
def process_payment(request, id):

    bill = get_object_or_404(
        Billing,
        id=id,
        status="unpaid"
    )

    if hasattr(bill, "payment"):
        messages.info(request, "This bill has already been paid.")
        return redirect("payment")

    if request.method == "POST":

        amount_paid = Decimal(request.POST.get("amount_paid"))

        if amount_paid < bill.total_amount:
            messages.error(request, "Insufficient payment.")
            return redirect("process_payment", id=id)

        Payment.objects.create(
            billing=bill,
            amount_paid=amount_paid,
            received_by=request.user,
            remarks=request.POST.get("remarks")
        )

        messages.success(request, "Payment processed successfully.")

        return redirect("payments")

    return render(
        request,
        "admin/process_payment.html",
        {
            "bill": bill
        }
    )



@login_required
def reports(request):

    payments = (
        Payment.objects
        .select_related(
            "billing",
            "billing__customer",
            "received_by"
        )
        .order_by("-payment_date")
    )

    # ===============================
    # Dashboard Summary
    # ===============================
    total_collections = payments.aggregate(
        total=Sum("amount_paid")
    )["total"] or Decimal("0.00")

    total_receipts = payments.count()

    total_consumption = Decimal("0.00")

    for payment in payments:
        total_consumption += payment.billing.consumption

    # ===============================
    # Consumption per Barangay
    # (Extracted from address)
    # ===============================
    barangay_totals = defaultdict(float)

    billings = Billing.objects.select_related("customer")

    for billing in billings:

        address = billing.customer.address or ""

        # Example:
        # Purok 1, Barangay 1, Pioduran, Albay
        parts = [part.strip() for part in address.split(",")]

        if len(parts) >= 2:
            barangay = parts[1]
        else:
            barangay = "Unknown"

        barangay_totals[barangay] += float(billing.consumption)

    barangay_labels = list(barangay_totals.keys())
    barangay_values = list(barangay_totals.values())

    # ===============================
    # Paid vs Unpaid Bills
    # ===============================
    paid = Billing.objects.filter(status="paid").count()
    unpaid = Billing.objects.filter(status="unpaid").count()

    # ===============================
    # Context
    # ===============================
    context = {
        "payments": payments,

        "total_collections": total_collections,
        "total_receipts": total_receipts,
        "total_consumption": total_consumption,

        "barangay_labels": json.dumps(barangay_labels),
        "barangay_values": json.dumps(barangay_values),

        "paid": paid,
        "unpaid": unpaid,
    }

    return render(request, "admin/reports.html", context)

@login_required
def official_receipt(request, payment_id):

    payment = get_object_or_404(
        Payment.objects.select_related(
            "billing",
            "billing__customer",
            "received_by"
        ),
        id=payment_id
    )

    return render(
        request,
        "admin/official_receipt.html",
        {
            "payment": payment
        }
    )

@login_required
def post_notification(request):

    customers = Customer.objects.order_by("fullname")
    notifications = Notification.objects.select_related("customer").order_by("-created_at")

    # Get unique barangays from customer addresses
    barangays = sorted({
        customer.address.split(",")[1].strip()
        for customer in customers
        if len(customer.address.split(",")) >= 2
    })

    if request.method == "POST":

        target = request.POST.get("target")
        customer_id = request.POST.get("customer")
        barangay = request.POST.get("barangay")
        status = request.POST.get("status")
        title = request.POST.get("title")
        message = request.POST.get("message")

        customer = None

        if target == "customer":
            if not customer_id:
                messages.error(request, "Please select a customer.")
                return redirect("post-notifacation")

            customer = Customer.objects.get(id=customer_id)

        elif target == "barangay":
            if not barangay:
                messages.error(request, "Please select a barangay.")
                return redirect("post-notifacation")

        Notification.objects.create(
            target=target,
            customer=customer,
            barangay=barangay if target == "barangay" else None,
            status=status,
            title=title,
            message=message,
        )

        messages.success(request, "Notification posted successfully.")
        return redirect("post-notifacation")

    return render(request, "admin/post_notification.html", {
        "customers": customers,
        "barangays": barangays,
        "notifications": notifications,
    })

@login_required
def delete_notification(request, pk):
    notification = get_object_or_404(Notification, pk=pk)
    notification.delete()

    messages.success(request, "Notification deleted successfully.")
    return redirect("post-notifacation")

@login_required
def paid_report(request):

    billings = (
        Billing.objects
        .filter(status="paid")
        .select_related("customer")
        .order_by("customer__fullname")
    )

    return render(request, "admin/paid_report.html", {
        "billings": billings
    })


@login_required
def unpaid_report(request):

    billings = (
        Billing.objects
        .filter(status="unpaid")
        .select_related("customer")
        .order_by("customer__fullname")
    )

    return render(request, "admin/unpaid_report.html", {
        "billings": billings
    })